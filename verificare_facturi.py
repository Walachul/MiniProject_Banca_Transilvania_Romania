from pathlib import Path
from PIL import Image, ImageOps
import pytesseract
from pdf2image import convert_from_path
import PyPDF2
import openpyxl
import os
import cv2
import re
import smtplib, ssl
"""
Sources:
Tesseract https://github.com/UB-Mannheim/tesseract/wiki


"""
# Store name of the folder containing invoices and
# other project files
checkInvDirPath = 'Proiect verificare facturi'
# variable storing a list of files in Proiect verificare facturi folder
listOfFilesInDir = os.listdir(checkInvDirPath)

# Variable for the name of the folder that will contain the jpeg invoices
nameFacturiJpeg = "facturi_jpeg"


# image size to be used for resizing of invoices
newInvoiceImageSize = (800,800)

try:
    # Open Excel Workbook named Baza_clienti with openpyxl package
    bazaClienti = openpyxl.load_workbook('Proiect verificare facturi/Baza_clienti.xlsx')
   # Open Sheet1
    sheet1 = bazaClienti['Sheet1']

    # Open Excel Workbook named Rezultat_VerificareFacturi_10.10.10
    verificareFacturi = openpyxl.load_workbook('Proiect verificare facturi/Rezultat_VerificareFacturi_10.10.10.xlsx')
    sheetVerificareFacturi = verificareFacturi['Sheet1']
    # Check if the first cell of column 2 has CUI value
    if sheet1['B1'].value == "CUI":
        
        # list that will keep the name of the invoices
        listOfJpegInvoices = []
        # get values of cells from CUI column
        # convert them to list
        listB = list(sheet1.columns)[1]
        
        
        # Select column B and iterate through cells
        for cell in listB[1:len(listB)]:
            
            # Read cell of column B starting with the second row
            CUI = cell.value
            # pdf invoice based on CUI
            pdfInvCUI = str(CUI)+'.pdf'
            # Check rotation of page

            # verify if pdf invoice file exists in the folder
            if pdfInvCUI in listOfFilesInDir:
                # use pdf to image package to convert pdf to image
                # this is an iterable object
                # need to use for loop
                pages = convert_from_path(checkInvDirPath+"/"+pdfInvCUI, 500)
                for page in pages:
                    # Create jpeg invoice image for tesseract reading
                    imageInvoice = str(CUI)+".jpg"
                    
                    # Add the name to listOfJpegInvoices
                    listOfJpegInvoices.append(imageInvoice)
                    # Save jpeg invoices in the facturi_jpeg folder
                    page.save(nameFacturiJpeg+"/"+imageInvoice, "JPEG")
         
        # variable storing a list of files in facturi_jpeg folder
        facturiJpegDir = os.listdir(nameFacturiJpeg)
        # Helper function for resizing of images
        for invoiceJpeg in facturiJpegDir:
            
            # Open the new invoice image with Pillow
            imgInvoice = Image.open(nameFacturiJpeg+"/"+invoiceJpeg)
            # With help of ImageOPs package
            # Check for EXIF orientation tag if image is 180 or other degree
            # If present, rotate the image accordingly
            imgInvoice = ImageOps.exif_transpose(imgInvoice)
           
           
            # check if width > 1000
            # apply new size of 800 X 800
            
            if imgInvoice.width > 1000:
                # resize 
                resizedInvoice = imgInvoice.resize(newInvoiceImageSize)
                # save and replace
                resizedInvoice.save(nameFacturiJpeg+"/"+invoiceJpeg)

        # Row counter 
        ROW = 2
       
        # ROW counter used for the Observatii column in rezultat_verificareFacturi workbook
        i = 3
        # List containing ROI for one invoice that needs to be compared
        listROIperInvoice = []
        # This reads each jpeg invoice
        # Then uses the OpenCV package to get the data
        for facturaJpeg in facturiJpegDir:
            print("Aceasta este factura jpeg: " + str(facturaJpeg))
            # Use Open CV package to read each jpeg invoice and
            # create rectangle zones / region of interest to get required information
            imgToRead = cv2.imread(nameFacturiJpeg+"/"+facturaJpeg)
            client = cv2.rectangle(imgToRead, (455, 88), (555,124),(0,255,0), 2)
            nrSiData = cv2.rectangle(imgToRead, (205, 276),(367,299),(0,255,0), 2)
            codProdus = cv2.rectangle(imgToRead, (18,331),(394,487),(0,255,0), 2)
            totalDePlata = cv2.rectangle(imgToRead, (527,477),(790,505),(0,255,0), 2)
            semnatura = cv2.rectangle(imgToRead, (392,518),(705,691),(0,255,0), 2)

            # Use Numpy slicing to crop the image and get the required information
            # this is going to be inserted in Tesseract to get text
            clientROI = client[88:124, 455:555]
            nrSiDataROI = nrSiData[276:299, 205:367]
            codProdusROI = codProdus[331:487, 18:394]
            totalDePlataROI = totalDePlata[477:505, 527:790]
            semnaturaROI = semnatura[518:691, 392:705]
            
            # Get data from nrSiDataROI with Tesseract
            # Split the string with separator "data:"
            # Needed as the nr of factura and data of invoice are located
            # in different columns in baza_clienti
            nrSiDataROI_DATA = pytesseract.image_to_string(nrSiDataROI).strip().split("data: ", maxsplit=1)
            #print(nrSiDataROI_DATA)
            # Append needed ROI to listROIperInvoice
            listROIperInvoice.extend((clientROI, nrSiDataROI_DATA[0],
                                    nrSiDataROI_DATA[1], totalDePlataROI))
            # display the image with ROI        
            cv2.imshow("image", client)
            
            
            # List of columns numbers from the baza client workbook
            # The column values needs to be compared with data extracted from invoices
            # This is an iterator object
            listCheckColumnValues = [1,3,4,6]
             # Counter i to help increment the column values
            j = 0
            # Column from listCheckColumnValues 
            COL = listCheckColumnValues[j]
             # cell value from baza_clienti based on row and col
            cellValue = str(sheet1.cell(row=ROW, column=COL).value).strip()
            # Compare values as per requirement
            for ROI in listROIperInvoice:
                # cell value from baza_clienti based on row and col
                # cellValue = str(sheet1.cell(row=ROW, column=COL).value).strip()
                # image is already processed 
                if type(ROI) is str:
                    ROI_DATA = ROI
                else:
                    # Get data from ROI with with Tesseract
                    ROI_DATA = str(pytesseract.image_to_string(ROI)).strip()
                print("Data from ROI: \n"+ ROI_DATA +"\n")
                
                print("Value of present cell: \n"+ cellValue+"\n")
                # Check if the value from the baza_clienti exists in
                # the data extracted with Tesseract
                if ROI_DATA.__contains__(cellValue):
                    print(True)

                    # Move to the next Column
                    # cell value
                    j+=1
                    if j == len(listCheckColumnValues):
                        
                        break
                    COL = listCheckColumnValues[j]
                    cellValue = str(sheet1.cell(row=ROW, column=COL).value).strip()
                                                            
                # Write    
                else:
                    print(False)
                    
                    sheetVerificareFacturi.cell(row=i, column=9).value = str(cellValue) + " diferit(a) pe factura " + str(facturaJpeg).split('.jpg')[0]
                    i+=1
                    
                    verificareFacturi.save("Proiect verificare facturi/Rezultat_VerificareFacturi_10.10.10.xlsx")
                    # Move to the next column
                    # cell value
                    j+=1
                    if j == len(listCheckColumnValues):
                        
                        break
                    COL = listCheckColumnValues[j]
                    cellValue = str(sheet1.cell(row=ROW, column=COL).value).strip()
                  
            
            # SECTION FOR DETECTING CODURILE PENTRU PRODUSE
            # AND ADDING THE CODES TO Baza_client workbook, col Cod_produs

            codProdusDATA = pytesseract.image_to_string(codProdusROI)
            getOnlyProductCodes = re.findall(r'[0-9]{3,20}', codProdusDATA)
            print("####### COD PRODUS(E) ######\n")
            print(getOnlyProductCodes)
            # Variable to concatenate two or more product codes
            longStringProductCodes = ""
            separator = ";"
            if getOnlyProductCodes != []:
                if len(getOnlyProductCodes) > 1:
                    longStringProductCodes = separator.join(getOnlyProductCodes)
                    sheet1.cell(row=ROW, column=5).value = longStringProductCodes
                    bazaClienti.save("Proiect verificare facturi/Baza_clienti.xlsx")
                else:
                    sheet1.cell(row=ROW, column=5).value = separator.join(getOnlyProductCodes)
                    bazaClienti.save("Proiect verificare facturi/Baza_clienti.xlsx")


            # SECTION FOR DETECTING IF SIGNATURE IS PRESENT
            # Currently it just gets the data
            # Hard and time constraining solution to be implement with Python
            semnaturaDATA = pytesseract.image_to_string(semnaturaROI)
            
            # wait for 2 seconds then destroy it
            cv2.waitKey(2000)  
            cv2.destroyAllWindows()

            # reset the list containing ROI data
            listROIperInvoice.clear()
            # Reset the j counter back to zero
            j = 0
            COL = listCheckColumnValues[j]
             # Move to the next row for the next invoice jpeg
            ROW+=1
           
            # increase the counter i for rezultat_verificarefacturi
            i+=1


    # SECTION FOR SENDING EMAIL
    # SSL PORT
    port = 465
    # Email address and password are stored in Environment Variables
    # This is for sender
    EMAIL_ADDRESS = os.environ.get('EMAIL_USER') 
    EMAIL_PASSWORD = os.environ.get('EMAIL_PASS') 
    
    # Receiver address
    EMAIL_RECEIVER = os.environ.get('EMAIL_ADDRESS_RECEIVER')
    message = """\
    Verificare facturi


    Buna ziua, 

    Procesul a fost finalizat cu succes. Se poate verifica rezultatul accesand sharefolderul dedicat.

    Spor,

    """
    # create a secure SLL Context
    contextSSL = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", port, context = contextSSL) as server:
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        print(">>> SENDING EMAIL >>>")
        server.sendmail(EMAIL_ADDRESS, EMAIL_RECEIVER, message)

    
except FileNotFoundError:
    print("Te rog verifică dacă fișierul există\nsau dacă numele este Baza_clienti.xlsx")
except PermissionError:
    print("Te rog închide fișierul Excel numit - Rezultat_VerificareFacturi")



